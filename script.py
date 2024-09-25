def process_excel_file(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Get the sheets
    sheet1 = wb["Lists"]
    sheet2 = wb["Links"]

    # Create a new sheet for the results
    result_sheet = wb.create_sheet("Results")

    # Process Sheet1 to create a list of unique names
    unique_names = []
    for cell in sheet1["A"]:
        if cell.value:
            names = [name.strip() for name in cell.value.split(",")]
            for name in names:
                if name not in unique_names:
                    unique_names.append(name)

    print(f"Number of unique names: {len(unique_names)}")
    print(f"Unique names: {unique_names}")

    # Extract links from Sheet2
    links = []
    for cell in sheet2["A"]:
        if cell.value:
            link = cell.value
            if not link.startswith("https://"):
                link = "https://" + link
            links.append(link)

    print(f"Number of links: {len(links)}")
    print(f"Links: {links}")

    # Create the key-value table in the Results sheet
    result_sheet["A1"] = "Name"
    result_sheet["B1"] = "Link"
    for i, (name, link) in enumerate(zip(unique_names, links), start=2):
        result_sheet[f"A{i}"] = name
        result_sheet[f"B{i}"] = link
        print(f"Writing to key-value table: {name} - {link}")

    # Process Sheet1 and create hyperlinks
    row_offset = len(unique_names) + 3  # Start after the key-value table
    result_sheet[f"A{row_offset}"] = "Updated Sheet1 with Hyperlinks"
    row_offset += 1

    for row_idx, row in enumerate(
        sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=1),
        start=row_offset,
    ):
        cell = row[0]
        if cell.value:
            names = [name.strip() for name in cell.value.split(",")]
            print(f"Processing names: {names}")
            for col_idx, name in enumerate(names, start=1):
                result_sheet.cell(row=row_idx, column=col_idx, value=name)
                if name in unique_names:
                    link_index = unique_names.index(name)
                    if link_index < len(links):
                        result_sheet.cell(row=row_idx, column=col_idx).hyperlink = (
                            links[link_index]
                        )
                        result_sheet.cell(row=row_idx, column=col_idx).style = (
                            "Hyperlink"
                        )
                        print(f"Added hyperlink for {name}: {links[link_index]}")
                    else:
                        print(f"No link available for {name}")
                else:
                    print(f"Name not found in unique names: {name}")

    # Save the workbook with a new filename
    new_file_path = file_path.replace(".xlsx", "_updated.xlsx")
    wb.save(new_file_path)
    print(f"File saved as: {new_file_path}")


# Usage
import os
import openpyxl

file_path = "/Users/hy/Code/DavidExcel/SampleData.xlsx"
try:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")

    process_excel_file(file_path)
    print("Excel file processed successfully.")
except Exception as e:
    print(f"An error occurred: {str(e)}")
